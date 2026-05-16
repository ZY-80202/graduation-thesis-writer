from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Dict, List, Sequence

from openpyxl import load_workbook

from thesis_skill.models import BackendEndpoint, DatabaseColumn, DatabaseTable, FrontendPage, ProjectProfile, ScreenshotAsset
from thesis_skill.parsers.docx_parser import read_docx_text
from thesis_skill.utils.file_utils import ensure_dir, read_text_file, relative_to, write_json
from thesis_skill.utils.image_utils import scan_images
from thesis_skill.utils.text_utils import compact_keywords, normalize_whitespace

TEXT_SUFFIXES = {".txt", ".md", ".rst"}
DOCX_SUFFIXES = {".docx"}
SPREADSHEET_SUFFIXES = {".xlsx", ".xlsm"}
CSV_SUFFIXES = {".csv", ".tsv"}
SQL_SUFFIXES = {".sql"}
CODE_SUFFIXES = {
    ".py",
    ".js",
    ".jsx",
    ".ts",
    ".tsx",
    ".vue",
    ".java",
    ".cs",
    ".php",
    ".go",
    ".html",
}
FRONTEND_SUFFIXES = {".vue", ".jsx", ".tsx", ".html", ".svelte"}
BACKEND_SUFFIXES = {".py", ".js", ".ts", ".java", ".cs", ".php", ".go"}


def parse_project(project_path: str | Path, output_dir: str | Path = "outputs/profiles") -> ProjectProfile:
    root = Path(project_path)
    if not root.exists() or not root.is_dir():
        raise FileNotFoundError(f"项目资料目录不存在: {root}")

    all_files = [item for item in root.rglob("*") if item.is_file()]
    source_materials = _collect_source_materials(root, all_files)
    db_tables = _parse_database_tables(root, all_files)
    endpoints = _parse_backend_endpoints(root, all_files)
    page_details = _detect_frontend_page_details(root, all_files)
    pages = [page.name for page in page_details]
    tech_stack = _detect_technology_stack(root, all_files, source_materials)
    modules = _detect_function_modules(root, source_materials, pages, endpoints, db_tables)
    roles = _detect_roles(source_materials)
    flows = _detect_business_flows(source_materials, modules)
    tests = _detect_test_points(source_materials, all_files, modules)
    screenshot_assets = _classify_screenshots(root, scan_images(root), modules)
    project_name = _detect_project_name(root, source_materials)

    profile = ProjectProfile(
        project_path=str(root),
        project_name=project_name,
        technology_stack=compact_keywords(tech_stack, 20),
        function_modules=compact_keywords(modules, 30),
        user_roles=compact_keywords(roles, 10),
        database_tables=db_tables,
        business_flows=compact_keywords(flows, 20),
        frontend_pages=compact_keywords(pages, 60),
        frontend_page_details=page_details[:200],
        backend_endpoints=endpoints[:300],
        test_points=compact_keywords(tests, 40),
        screenshots=[asset.path for asset in screenshot_assets[:300]],
        screenshot_assets=screenshot_assets[:300],
        source_materials=source_materials,
        warnings=[],
    )

    if not profile.function_modules:
        profile.warnings.append("未能从项目资料中识别明确功能模块，生成正文时会保留待补充标记。")
    if not profile.database_tables:
        profile.warnings.append("未识别到数据库表结构，如有 SQL 或表结构文档请补充到项目目录。")
    if not profile.backend_endpoints:
        profile.warnings.append("未识别到后端接口，请检查 source_code/backend/ 或接口文档。")
    if not profile.frontend_page_details:
        profile.warnings.append("未识别到前端页面，请检查 source_code/frontend/、pages/ 或 views/ 目录。")
    ensure_dir(output_dir)
    write_json(Path(output_dir) / "project_profile.json", profile)
    return profile


def _collect_source_materials(root: Path, files: Sequence[Path]) -> Dict[str, str]:
    materials: Dict[str, str] = {}
    for file_path in files:
        suffix = file_path.suffix.lower()
        relative = relative_to(file_path, root)
        try:
            if suffix in TEXT_SUFFIXES:
                materials[relative] = read_text_file(file_path, max_chars=12000)
            elif suffix in DOCX_SUFFIXES:
                materials[relative] = read_docx_text(file_path)[:12000]
            elif suffix in CSV_SUFFIXES:
                materials[relative] = _read_csv_summary(file_path)
            elif suffix in SPREADSHEET_SUFFIXES:
                materials[relative] = _read_xlsx_summary(file_path)
            elif suffix in SQL_SUFFIXES:
                materials[relative] = read_text_file(file_path, max_chars=60000)
        except Exception as exc:
            materials[relative] = f"[读取失败] {exc}"
    return materials


def _read_csv_summary(path: Path) -> str:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    rows: List[str] = []
    with path.open("r", encoding="utf-8-sig", errors="ignore", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        for index, row in enumerate(reader):
            rows.append(" | ".join(row))
            if index >= 30:
                break
    return "\n".join(rows)


def _read_xlsx_summary(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    parts: List[str] = []
    for sheet_name in workbook.sheetnames[:8]:
        sheet = workbook[sheet_name]
        parts.append(f"Sheet: {sheet_name}")
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            values = ["" if cell is None else str(cell) for cell in row]
            parts.append(" | ".join(values))
            if index >= 20:
                break
    workbook.close()
    return "\n".join(parts)


def _detect_technology_stack(root: Path, files: Sequence[Path], materials: Dict[str, str]) -> List[str]:
    stack: List[str] = []
    file_names = {file.name.lower(): file for file in files}
    suffixes = {file.suffix.lower() for file in files}

    if "package.json" in file_names:
        stack.extend(["Node.js", "JavaScript/TypeScript"])
        try:
            data = json.loads(read_text_file(file_names["package.json"]))
            deps = {**data.get("dependencies", {}), **data.get("devDependencies", {})}
            for key in deps:
                name = key.lower()
                if name in {"vue", "react", "angular", "vite", "webpack", "element-plus", "antd", "axios", "express", "koa", "nestjs"}:
                    stack.append(key)
        except Exception:
            pass
    if "requirements.txt" in file_names or "pyproject.toml" in file_names or ".py" in suffixes:
        stack.append("Python")
    if "pom.xml" in file_names or ".java" in suffixes:
        stack.extend(["Java", "Spring Boot"])
    if any(file.name.endswith(".csproj") for file in files) or ".cs" in suffixes:
        stack.extend(["C#", ".NET"])
    if "go.mod" in file_names or ".go" in suffixes:
        stack.append("Go")
    if ".vue" in suffixes:
        stack.append("Vue")
    if ".tsx" in suffixes or ".jsx" in suffixes:
        stack.append("React")
    text = "\n".join(materials.values()).lower()
    for database in ["mysql", "postgresql", "postgres", "sqlite", "redis", "mongodb", "sql server"]:
        if database in text:
            stack.append({"postgres": "PostgreSQL", "mysql": "MySQL"}.get(database, database.title()))
    if any(file.suffix.lower() == ".sql" for file in files):
        stack.append("关系型数据库")
    return stack


def _parse_database_tables(root: Path, files: Sequence[Path]) -> List[DatabaseTable]:
    tables: Dict[str, DatabaseTable] = {}
    for file_path in files:
        if file_path.suffix.lower() != ".sql":
            continue
        text = read_text_file(file_path, max_chars=120000)
        for table_name, body, tail in _iter_create_table_blocks(text):
            columns, primary_keys = _parse_table_body(body)
            for column in columns:
                if column.name in primary_keys:
                    column.is_primary_key = True
                    column.nullable = False
            table_comment = _extract_table_comment(tail)
            tables[table_name] = DatabaseTable(
                name=table_name,
                columns=[column.name for column in columns],
                field_details=columns,
                comment=table_comment,
                source=relative_to(file_path, root),
            )
        for match in re.finditer(r"(?:insert\s+into|alter\s+table)\s+[`\"\[]?([\w\u4e00-\u9fff.]+)[`\"\]]?", text, re.I):
            name = _clean_sql_name(match.group(1))
            tables.setdefault(name, DatabaseTable(name=name, columns=[], field_details=[], source=relative_to(file_path, root)))
    return list(tables.values())


def _iter_create_table_blocks(sql: str) -> List[tuple[str, str, str]]:
    blocks: List[tuple[str, str, str]] = []
    pattern = re.compile(r"create\s+table\s+(?:if\s+not\s+exists\s+)?([`\"\[]?)([\w\u4e00-\u9fff.]+)[`\"\]]?\s*\(", re.I)
    for match in pattern.finditer(sql):
        name = _clean_sql_name(match.group(2))
        open_index = sql.find("(", match.end() - 1)
        close_index = _find_matching_paren(sql, open_index)
        if close_index == -1:
            continue
        semicolon = sql.find(";", close_index)
        if semicolon == -1:
            semicolon = min(len(sql), close_index + 400)
        body = sql[open_index + 1 : close_index]
        tail = sql[close_index + 1 : semicolon]
        blocks.append((name, body, tail))
    return blocks


def _find_matching_paren(text: str, open_index: int) -> int:
    depth = 0
    quote = ""
    index = open_index
    while index < len(text):
        char = text[index]
        if quote:
            if char == quote and text[index - 1] != "\\":
                quote = ""
        elif char in {"'", '"'}:
            quote = char
        elif char == "(":
            depth += 1
        elif char == ")":
            depth -= 1
            if depth == 0:
                return index
        index += 1
    return -1


def _parse_table_body(body: str) -> tuple[List[DatabaseColumn], set[str]]:
    columns: List[DatabaseColumn] = []
    primary_keys: set[str] = set()
    for item in _split_sql_items(body):
        definition = item.strip().rstrip(",")
        lower = definition.lower()
        if not definition:
            continue
        pk_match = re.search(r"primary\s+key\s*\(([^)]+)\)", definition, re.I)
        if pk_match:
            primary_keys.update(_clean_sql_name(value) for value in pk_match.group(1).split(","))
        if lower.startswith(("primary ", "key ", "unique ", "index ", "constraint ", "foreign ")):
            continue
        column = _parse_column_definition(definition)
        if column:
            columns.append(column)
    return columns, primary_keys


def _split_sql_items(body: str) -> List[str]:
    items: List[str] = []
    depth = 0
    quote = ""
    start = 0
    for index, char in enumerate(body):
        if quote:
            if char == quote and body[index - 1] != "\\":
                quote = ""
        elif char in {"'", '"'}:
            quote = char
        elif char == "(":
            depth += 1
        elif char == ")":
            depth -= 1
        elif char == "," and depth == 0:
            items.append(body[start:index])
            start = index + 1
    items.append(body[start:])
    return items


def _parse_column_definition(definition: str) -> DatabaseColumn | None:
    match = re.match(r"\s*[`\"\[]?([\w\u4e00-\u9fff-]+)[`\"\]]?\s+(.+)$", definition, re.S)
    if not match:
        return None
    name = match.group(1)
    rest = re.sub(r"\s+", " ", match.group(2).strip())
    lower = rest.lower()
    constraint_words = {
        "not",
        "null",
        "default",
        "comment",
        "primary",
        "unique",
        "key",
        "auto_increment",
        "references",
        "constraint",
        "collate",
        "character",
        "check",
        "generated",
    }
    tokens = _split_type_tokens(rest)
    type_parts: List[str] = []
    for token in tokens:
        if token.lower() in constraint_words:
            break
        type_parts.append(token)
    data_type = " ".join(type_parts).strip()
    comment = _extract_column_comment(rest)
    default = _extract_default(rest)
    is_pk = "primary key" in lower
    nullable = not ("not null" in lower or is_pk)
    return DatabaseColumn(name=name, data_type=data_type, is_primary_key=is_pk, nullable=nullable, default=default, comment=comment)


def _split_type_tokens(text: str) -> List[str]:
    tokens: List[str] = []
    current: List[str] = []
    depth = 0
    for char in text:
        if char == "(":
            depth += 1
        elif char == ")":
            depth -= 1
        if char.isspace() and depth == 0:
            if current:
                tokens.append("".join(current))
                current = []
        else:
            current.append(char)
    if current:
        tokens.append("".join(current))
    return tokens


def _extract_column_comment(text: str) -> str:
    match = re.search(r"comment\s+'([^']*)'", text, re.I)
    if match:
        return match.group(1)
    match = re.search(r'comment\s+"([^"]*)"', text, re.I)
    return match.group(1) if match else ""


def _extract_default(text: str) -> str:
    match = re.search(r"\bdefault\s+((?:'[^']*')|(?:\"[^\"]*\")|[^\s,]+)", text, re.I)
    return match.group(1).strip("'\"") if match else ""


def _extract_table_comment(text: str) -> str:
    match = re.search(r"comment\s*=\s*'([^']*)'", text, re.I)
    if match:
        return match.group(1)
    match = re.search(r'comment\s*=\s*"([^"]*)"', text, re.I)
    return match.group(1) if match else ""


def _clean_sql_name(value: str) -> str:
    clean = value.strip().strip("`\"[]")
    if "." in clean:
        clean = clean.split(".")[-1]
    return clean.strip("`\"[] ")


def _parse_backend_endpoints(root: Path, files: Sequence[Path]) -> List[BackendEndpoint]:
    endpoints: List[BackendEndpoint] = []
    backend_files = [file for file in files if _is_backend_route_file(root, file)]
    if not backend_files:
        backend_files = [file for file in files if file.suffix.lower() in BACKEND_SUFFIXES]
    for file_path in backend_files:
        try:
            text = read_text_file(file_path, max_chars=100000)
        except Exception:
            continue
        endpoints.extend(_extract_endpoints_from_text(root, file_path, text))
    unique: Dict[str, BackendEndpoint] = {}
    for endpoint in endpoints:
        key = f"{endpoint.method.upper()}:{endpoint.path}"
        unique.setdefault(key, endpoint)
    return list(unique.values())


def _is_backend_route_file(root: Path, file_path: Path) -> bool:
    if file_path.suffix.lower() not in BACKEND_SUFFIXES:
        return False
    relative = Path(relative_to(file_path, root))
    parts = [part.lower() for part in relative.parts]
    name = file_path.stem.lower()
    return (
        ("source_code" in parts and "backend" in parts)
        or "backend" in parts
        or "routes" in parts
        or "controllers" in parts
        or "controller" in name
        or "route" in name
        or "api" in name
    )


def _extract_endpoints_from_text(root: Path, file_path: Path, text: str) -> List[BackendEndpoint]:
    source = relative_to(file_path, root)
    endpoints: List[BackendEndpoint] = []
    spring_base = _first_match(text, r"@RequestMapping\s*\(\s*[\"']([^\"']+)[\"']")
    nest_base = _first_match(text, r"@Controller\s*\(\s*[\"']?([^\"')]+)?[\"']?\s*\)") or ""
    base_path = spring_base
    if not base_path and nest_base:
        base_path = "/" + nest_base.strip("/")

    patterns = [
        ("spring", re.compile(r"@(GetMapping|PostMapping|PutMapping|DeleteMapping|PatchMapping|RequestMapping)\s*(?:\(\s*(?:value\s*=\s*)?[\"']([^\"']*)[\"']|\([^)]*method\s*=\s*RequestMethod\.(GET|POST|PUT|DELETE|PATCH)[^)]*)?", re.I)),
        ("express", re.compile(r"\b(?:router|app)\.(get|post|put|delete|patch|all)\s*\(\s*[\"'`]([^\"'`]+)[\"'`]\s*,?\s*([A-Za-z_]\w*)?", re.I)),
        ("fastapi", re.compile(r"@(?:app|router)\.(get|post|put|delete|patch)\s*\(\s*[\"']([^\"']+)[\"']", re.I)),
        ("flask", re.compile(r"@(?:app|bp|blueprint)\.route\s*\(\s*[\"']([^\"']+)[\"'](?:,\s*methods\s*=\s*\[([^\]]+)\])?", re.I)),
        ("nestjs", re.compile(r"@(Get|Post|Put|Delete|Patch)\s*\(\s*[\"']?([^\"')]+)?[\"']?\s*\)", re.I)),
        ("plain", re.compile(r"\b(GET|POST|PUT|DELETE|PATCH)\s+(/[^\s,;]+)", re.I)),
    ]
    for kind, pattern in patterns:
        for match in pattern.finditer(text):
            method = "GET"
            path = ""
            name = ""
            if kind == "spring":
                mapping = (match.group(1) or "RequestMapping").lower()
                method = (match.group(3) or mapping.replace("mapping", "") or "ANY").upper()
                path = match.group(2) or ""
                path = _join_paths(base_path, path)
            elif kind == "flask":
                path = match.group(1)
                methods_raw = match.group(2)
                method = _first_http_method(methods_raw) if methods_raw else "GET"
            elif kind == "nestjs":
                method = match.group(1).upper()
                path = _join_paths(base_path, match.group(2) or "")
            elif kind == "plain":
                method = match.group(1).upper()
                path = match.group(2)
            else:
                method = match.group(1).upper()
                path = match.group(2)
                if kind == "express" and match.lastindex and match.lastindex >= 3:
                    name = match.group(3) or ""
            if not path:
                path = "/"
            endpoints.append(
                BackendEndpoint(
                    method=method,
                    path=path,
                    name=name,
                    description=_endpoint_description(method, path),
                    source=source,
                )
            )
    return endpoints


def _first_match(text: str, pattern: str) -> str:
    match = re.search(pattern, text, re.I)
    return match.group(1) if match and match.group(1) else ""


def _first_http_method(raw: str | None) -> str:
    if not raw:
        return "GET"
    match = re.search(r"GET|POST|PUT|DELETE|PATCH", raw, re.I)
    return match.group(0).upper() if match else "GET"


def _join_paths(base: str, path: str) -> str:
    combined = "/".join(part.strip("/") for part in [base, path] if part is not None and part != "")
    return "/" + combined.strip("/") if combined.strip("/") else "/"


def _endpoint_description(method: str, path: str) -> str:
    module = _humanize_identifier(path.strip("/").split("/")[0] if path.strip("/") else "index")
    action = {
        "GET": "查询",
        "POST": "新增或提交",
        "PUT": "修改",
        "PATCH": "局部修改",
        "DELETE": "删除",
        "ANY": "处理",
        "ALL": "处理",
    }.get(method.upper(), "处理")
    return f"用于{action}{module}相关数据"


def _detect_frontend_page_details(root: Path, files: Sequence[Path]) -> List[FrontendPage]:
    pages: List[FrontendPage] = []
    frontend_files = [file for file in files if _is_frontend_page_file(root, file)]
    for file_path in frontend_files:
        try:
            text = read_text_file(file_path, max_chars=50000)
        except Exception:
            text = ""
        name = _humanize_identifier(file_path.stem)
        title = _extract_page_title(text) or name
        route = _extract_frontend_route(file_path, text)
        pages.append(
            FrontendPage(
                name=name,
                file_path=relative_to(file_path, root),
                route_path=route,
                title=title,
                matched_module=_module_from_text(f"{file_path.stem} {title} {route}"),
            )
        )
    unique: Dict[str, FrontendPage] = {}
    for page in pages:
        unique.setdefault(page.file_path, page)
    return list(unique.values())


def _is_frontend_page_file(root: Path, file_path: Path) -> bool:
    if file_path.suffix.lower() not in FRONTEND_SUFFIXES:
        return False
    relative = Path(relative_to(file_path, root))
    parts = [part.lower() for part in relative.parts]
    return (
        ("source_code" in parts and "frontend" in parts)
        or "frontend" in parts
        or "pages" in parts
        or "views" in parts
        or "templates" in parts
    )


def _extract_page_title(text: str) -> str:
    for pattern in [
        r"<title>([^<]+)</title>",
        r"<h1[^>]*>([^<]+)</h1>",
        r"title\s*[:=]\s*[\"']([^\"']+)[\"']",
        r"name\s*[:=]\s*[\"']([^\"']+)[\"']",
    ]:
        match = re.search(pattern, text, re.I)
        if match:
            return normalize_whitespace(match.group(1))[:40]
    return ""


def _extract_frontend_route(file_path: Path, text: str) -> str:
    match = re.search(r"path\s*[:=]\s*[\"']([^\"']+)[\"']", text)
    if match:
        return match.group(1)
    stem = file_path.stem
    if stem.lower() in {"index", "home"}:
        return "/"
    return "/" + re.sub(r"([a-z0-9])([A-Z])", r"\1-\2", stem).replace("_", "-").lower()


def _classify_screenshots(root: Path, image_paths: Sequence[Path], modules: Sequence[str]) -> List[ScreenshotAsset]:
    assets: List[ScreenshotAsset] = []
    for path in image_paths:
        relative = relative_to(path, root)
        text = f"{path.stem} {' '.join(path.parts)}"
        module = _module_from_text(text, modules)
        section = _section_from_screenshot_name(text, module)
        caption = _caption_from_screenshot(path.stem, module, section)
        assets.append(
            ScreenshotAsset(
                path=relative,
                file_name=path.name,
                inferred_section=section,
                matched_module=module,
                caption=caption,
            )
        )
    return assets


def _section_from_screenshot_name(text: str, module: str = "") -> str:
    lower = text.lower()
    if any(key in lower for key in ["login", "register", "auth", "signin", "登录", "注册"]):
        return "登录"
    if any(key in lower for key in ["admin", "dashboard", "manage", "后台", "管理"]):
        return "后台管理"
    if any(key in lower for key in ["test", "测试"]):
        return "系统测试"
    if any(key in lower for key in ["database", "db", "er", "数据库"]):
        return "数据库设计"
    return module or "系统详细设计与实现"


def _caption_from_screenshot(stem: str, module: str, section: str) -> str:
    readable = _humanize_identifier(stem)
    if readable and readable.lower() != stem.lower():
        return f"{readable}页面运行效果"
    if module:
        return f"{module}页面运行效果"
    return f"{section}页面运行效果"


def _detect_function_modules(
    root: Path,
    materials: Dict[str, str],
    pages: Sequence[str],
    endpoints: Sequence[BackendEndpoint],
    tables: Sequence[DatabaseTable],
) -> List[str]:
    candidates: List[str] = []
    haystack = "\n".join([*materials.keys(), *materials.values(), *pages, *[ep.path for ep in endpoints], *[t.name for t in tables]]).lower()
    keyword_map = {
        "登录注册": ["login", "register", "auth", "signin", "登录", "注册", "认证"],
        "用户管理": ["user", "member", "profile", "用户", "会员"],
        "商品管理": ["product", "goods", "item", "商品", "物品"],
        "订单管理": ["order", "订单"],
        "购物车": ["cart", "购物车"],
        "评论留言": ["comment", "message", "review", "留言", "评论"],
        "后台管理": ["admin", "dashboard", "manage", "后台", "管理端"],
        "数据统计": ["statistics", "report", "chart", "报表", "统计"],
        "文件上传": ["upload", "file", "上传", "附件"],
        "权限管理": ["permission", "role", "权限", "角色"],
    }
    for module, keys in keyword_map.items():
        if any(key.lower() in haystack for key in keys):
            candidates.append(module)
    for page in pages:
        value = _humanize_identifier(page)
        if value and len(value) <= 18:
            candidates.append(value)
    for endpoint in endpoints[:100]:
        parts = [part for part in endpoint.path.split("/") if part and not part.startswith("{") and not part.startswith(":")]
        if parts:
            candidates.append(_humanize_identifier(parts[0]))
    for table in tables:
        candidates.append(_humanize_identifier(table.name))
    return candidates


def _module_from_text(text: str, modules: Sequence[str] | None = None) -> str:
    modules = modules or []
    lower = text.lower()
    for module in modules:
        if module and (module.lower() in lower or any(token in lower for token in _module_tokens(module))):
            return module
    mapping = [
        ("登录注册", ["login", "register", "auth", "signin", "登录", "注册"]),
        ("用户管理", ["user", "member", "profile", "用户", "会员"]),
        ("商品管理", ["product", "goods", "item", "商品", "物品"]),
        ("订单管理", ["order", "订单"]),
        ("购物车", ["cart", "购物车"]),
        ("评论留言", ["comment", "message", "review", "留言", "评论"]),
        ("后台管理", ["admin", "dashboard", "manage", "后台"]),
        ("数据统计", ["statistics", "report", "chart", "统计", "报表"]),
        ("文件上传", ["upload", "file", "上传"]),
        ("权限管理", ["permission", "role", "权限", "角色"]),
    ]
    for module, keys in mapping:
        if any(key in lower for key in keys):
            return module
    return ""


def _module_tokens(module: str) -> List[str]:
    mapping = {
        "登录": ["login", "auth"],
        "注册": ["register"],
        "用户": ["user", "member"],
        "商品": ["product", "goods"],
        "订单": ["order"],
        "购物车": ["cart"],
        "评论": ["comment", "review"],
        "留言": ["message"],
        "后台": ["admin", "dashboard"],
        "统计": ["statistics", "report"],
    }
    tokens: List[str] = []
    for key, values in mapping.items():
        if key in module:
            tokens.extend(values)
    return tokens


def _detect_roles(materials: Dict[str, str]) -> List[str]:
    text = "\n".join(materials.values())
    roles = []
    for role in ["管理员", "普通用户", "游客", "教师", "学生", "商家", "会员", "审核员"]:
        if role in text:
            roles.append(role)
    if "管理员" not in roles and re.search(r"\badmin\b|后台", text, re.I):
        roles.append("管理员")
    if not roles:
        roles.extend(["普通用户", "管理员"])
    return roles


def _detect_business_flows(materials: Dict[str, str], modules: Sequence[str]) -> List[str]:
    text = "\n".join(materials.values())
    flows = re.findall(r"(?:流程|步骤|业务)[：:]\s*([^\n。；;]+)", text)
    if flows:
        return flows
    if any("订单" in item for item in modules):
        return ["用户浏览商品", "加入购物车", "提交订单", "支付或确认订单", "后台处理订单"]
    return ["用户登录系统", "选择功能模块", "提交业务数据", "系统校验并保存", "返回处理结果"]


def _detect_test_points(materials: Dict[str, str], files: Sequence[Path], modules: Sequence[str]) -> List[str]:
    tests = [module for module in modules if module]
    if any("test" in file.name.lower() or "测试" in file.name for file in files):
        tests.append("单元测试与功能测试")
    text = "\n".join(materials.values())
    for match in re.findall(r"(?:测试|用例)[：:]\s*([^\n。；;]+)", text):
        tests.append(match)
    return tests


def _detect_project_name(root: Path, materials: Dict[str, str]) -> str:
    for text in materials.values():
        for pattern in [r"项目名称[：:]\s*(.+)", r"系统名称[：:]\s*(.+)", r"#\s+(.+)"]:
            match = re.search(pattern, text)
            if match:
                return normalize_whitespace(match.group(1))[:40]
    return root.name


def _humanize_identifier(value: str) -> str:
    clean = re.sub(r"[_\-]+", " ", value)
    clean = re.sub(r"(?<!^)([A-Z])", r" \1", clean).strip()
    mapping = {
        "login": "登录",
        "signin": "登录",
        "auth": "登录认证",
        "register": "注册",
        "user": "用户",
        "member": "会员",
        "admin": "后台管理",
        "dashboard": "数据看板",
        "product": "商品",
        "goods": "商品",
        "item": "商品",
        "order": "订单",
        "cart": "购物车",
        "comment": "评论",
        "review": "评论",
        "message": "留言",
        "upload": "文件上传",
        "role": "角色",
        "permission": "权限",
    }
    lower = clean.lower()
    for key, zh in mapping.items():
        if key in lower:
            return zh
    return clean.strip().title()
